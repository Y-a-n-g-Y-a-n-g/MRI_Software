# -*- coding: utf-8 -*-
"""
@Time ： 12/23/2022 10:23 PM
@Auth ： YY
@File ：main.py.py
@IDE ：PyCharm
@state:
@Function：
"""
import os
import sys

import vtk
from PIL import Image
import cv2
import numpy as np
import torch
from PyQt5 import QtWidgets, QtCore
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QImage, QPixmap, QStandardItemModel, QStandardItem, QCursor
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QTableView, QMessageBox
from openpyxl.reader.excel import load_workbook

import MaximumView
import ViewLogs
from GUI import Ui_MainWindow
import SimpleITK as sitk
from vtk.qt.QVTKRenderWindowInteractor import QVTKRenderWindowInteractor

from model_Unet import UNet
from utlis import dicom_to_numpy, BasicDataset, put_mask, traversal_files, draw_axes, DrawaPlane, DrawAMarker, DrawProstateCenter, \
    DrawaPlanep1p2p3p4, MarkercheckAixs
from torchvision import transforms
import torch.nn.functional as F

current_directory = os.getcwd()


class MetaTableView(QTableView):
    def __init__(self, parent=None):
        super(MetaTableView, self).__init__(parent)
        self.resize(1300, 800)
        self.myModel = QStandardItemModel()  # model
        self.initHeader()  # 初始化表头
        self.setModel(self.myModel)
        # self.initData()  # 初始化模拟数据

    def initHeader(self):
        self.myModel.setHorizontalHeaderItem(0, QStandardItem("Tags"))
        self.myModel.setHorizontalHeaderItem(1, QStandardItem("value"))
        self.myModel.setHorizontalHeaderItem(2, QStandardItem("VR"))
        self.myModel.setHorizontalHeaderItem(3, QStandardItem("Name"))
        self.myModel.setHorizontalHeaderItem(4, QStandardItem("Retired State"))
        self.myModel.setHorizontalHeaderItem(5, QStandardItem("Tips"))
        self.myModel.setHorizontalHeaderItem(6, QStandardItem("中文备注-will delete later"))

    def initData(self, Data=None):
        for row in range(len(Data)):
            for col in range(len(Data[0])):
                if Data[row][col] != "None":
                    item = QStandardItem(Data[row][col])
                    item.setTextAlignment(Qt.AlignCenter)
                    self.myModel.setItem(row, col, item)

        self.horizontalHeader().resizeSection(0, 90)
        self.horizontalHeader().resizeSection(1, 300)
        self.horizontalHeader().resizeSection(2, 100)
        self.horizontalHeader().resizeSection(3, 200)
        self.horizontalHeader().resizeSection(4, 90)
        self.horizontalHeader().resizeSection(5, 90)
        self.horizontalHeader().resizeSection(6, 300)


from PyQt5.uic import loadUi


class ViewLogsWindow(QMainWindow, ViewLogs.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)


class ViewRYZWindow(QMainWindow, MaximumView.Ui_MainWindow):
    def __init__(self, windowname):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle(windowname)


# class MainWindow(QMainWindow):
#     def __init__(self, parent=None):
#         super(MainWindow, self).__init__(parent)
#         loadUi('GUI.ui', self)
class MyApp(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.logtagnumber = 0
        self.unet_init()
        self.view3D_init()
        self.dicomheadercheck_init()
        self.software_init()
        self.ViewLogsWindow = ViewLogsWindow()
        self.view_R_fullScreen = ViewRYZWindow('R-view')
        self.view_Y_fullScreen = ViewRYZWindow('Y-view')
        self.view_Z_fullScreen = ViewRYZWindow("Z-view")
        self.add_logs("System initialization is completed.")

    def software_init(self):
        self.horizontalSlider_R_view.setEnabled(False)
        self.horizontalSlider_Y_view.setEnabled(False)
        self.horizontalSlider_Z_view.setEnabled(False)
        self.image_array = None
        self.Rcorners = []
        self.Zcorners = []
        self.Ycorners = []
        self.log_history = []
        self.segementmask = []
        self.metadata = []
        self.Image_Position = []
        self.Image_Orientation = []
        self.Pixel_Spacing = []
        self.Patient_Position = []
        self.marker_position = {}
        self.marker_position['mark1'] = [0, 0, 0]
        self.marker_position['mark2'] = [0, 0, 0]
        self.marker_position['mark3'] = [0, 0, 0]
        self.marker_position['mark4'] = [0, 0, 0]
        self.cursor = QCursor(Qt.CrossCursor)
        self.current_slice_position_x = 0
        self.current_slice_position_y = 0
        self.current_slice_position_z = 0
        self.current_R_view_frame = 0
        self.current_Y_view_frame = 0
        self.current_Z_view_frame = 0

        self.flag_Image_Position_readSuccessed = True
        self.flag_Image_Orientation_readSuccessed = True
        self.flag_Pixel_Spacing_readSuccessed = True
        self.flag_Patient_Position_readSuccessed = True

    def unet_init(self):
        "Finished. No need to change it."
        self.device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
        # 加载模型load model
        self.net = UNet(n_channels=1, n_classes=2, bilinear=True)
        self.net.load_state_dict(torch.load(r'C:\Users\Yang\Desktop\Unet\dogprostate-TP+2NT+2NG_0.0001_epoch250.pth', map_location=self.device))
        self.net.to(device=self.device)

    def dicomheadercheck_init(self):
        "Finished. No need to change it."
        wb = load_workbook('data/DICOMheaderinformation.xlsx')
        sheet_ranges = wb['Sheet1']
        self.tags = {}
        for i in range(4, sheet_ranges.max_row + 1):
            self.tags[str(sheet_ranges['A' + str(i)].value)] = [str(sheet_ranges['B' + str(i)].value), str(sheet_ranges['C' + str(i)].value),
                                                                str(sheet_ranges['D' + str(i)].value), str(sheet_ranges['E' + str(i)].value),
                                                                str(sheet_ranges['F' + str(i)].value)]

    def view3D_init(self):

        self.vl = QtWidgets.QVBoxLayout()
        self.vtkWidget = QVTKRenderWindowInteractor(self.frame_3Dview)
        self.vl.addWidget(self.vtkWidget)
        self.ren = vtk.vtkRenderer()
        self.ren.SetBackground(0.4, 0.4, 0.4)
        self.vtkWidget.GetRenderWindow().AddRenderer(self.ren)
        self.iren = self.vtkWidget.GetRenderWindow().GetInteractor()

        self.ren.ResetCamera()
        self.frame_3Dview.setLayout(self.vl)
        self.iren.Initialize()
        self.update_ax()

    def f_btn_reset3dview(self):
        camera = vtk.vtkCamera()

        # 设置摄像机位置为 (0,0,1)，然后朝向 (0,0,0)
        camera.SetPosition(0, 0, 0)  # 摄像机的方向
        camera.SetFocalPoint(-15, -15, -15)  # 摄像机的朝向

        # 设置摄像机的上方向为 (0,1,0)
        camera.SetViewUp(0, 0, 1)

        # 设置摄像机视角范围（即可视距离）为 1
        camera.SetClippingRange(1, 100)

        # 重置视图
        self.ren.SetActiveCamera(camera)
        self.ren.ResetCamera()
        self.vtkWidget.GetRenderWindow().Render()
        self.add_logs("Reset the perspective of the 3D view.")

    def f_checkBox_ShowSegement(self, sigbool):
        # print(sigbool)
        if sigbool:
            self.add_logs("Will show the automatic segmentation of the prostate.")
        else:
            self.add_logs("Will hide the automatic segmentation of the prostate.")

    def f_btn_CopyTransformationT(self):
        self.add_logs("Sorry, this function is still under development!")

    def updateRPlane(self, xPosition, Translate):
        try:
            self.ren.RemoveActor(self.PlaneActor)
        except:
            pass
        self.PlaneActor = DrawaPlane(xPosition, Translate)

        self.ren.AddActor(self.PlaneActor)
        self.vtkWidget.GetRenderWindow().Render()

    def add_logs(self, logstr):
        log = f"#[{self.logtagnumber}]:{logstr}"
        self.log_history.append(log)
        self.label_singlelog.setText(log)
        self.logtagnumber = self.logtagnumber + 1

    def update_ax(self):
        AxesActor = draw_axes(Translate=(0, 0, 0), RotateX=0, RotateY=0, RotateZ=0, XYZLength=(10, 10, 10))
        self.ren.AddActor(AxesActor)
        self.vtkWidget.GetRenderWindow().Render()

    def mask_to_image(self, mask: np.ndarray):
        "Finished. No need to change it."
        if mask.ndim == 2:
            return Image.fromarray((mask * 255).astype(np.uint8))
        elif mask.ndim == 3:
            return Image.fromarray((np.argmax(mask, axis=0) * 255 / mask.shape[0]).astype(np.uint8))

    def predict_img(self, full_img, scale_factor=1.0, out_threshold=0.5):
        "Finished. No need to change it."
        self.net.eval()
        img = torch.from_numpy(BasicDataset.preprocess(full_img, scale_factor, is_mask=False))
        img = img.unsqueeze(0)
        img = img.to(device=self.device, dtype=torch.float32)
        with torch.no_grad():
            output = self.net(img)
            if self.net.n_classes > 1:
                probs = F.softmax(output, dim=1)[0]
            else:
                probs = torch.sigmoid(output)[0]
            tf = transforms.Compose([
                transforms.ToPILImage(),
                transforms.Resize((full_img.size[1], full_img.size[0])),
                transforms.ToTensor()
            ])
            full_mask = tf(probs.cpu()).squeeze()
        if self.net.n_classes == 1:
            return (full_mask > out_threshold).numpy()
        else:
            return F.one_hot(full_mask.argmax(dim=0), self.net.n_classes).permute(2, 0, 1).numpy()

    def f_btn_findmarker(self):
        self.add_logs(f'The automatic finding Marker is not available now!And it is still under development.')

    #肿瘤自动分割
    def f_btn_ProstateSegment(self):
        self.add_logs("Start to segment images,please wait...")
        a = QMessageBox.question(self, 'Information',
                                 'The existing prostate segmentation model is trained based on human prostate data sets. It is well tested on the image of the human prostate, but the results may not be too ideal on other pictures.',
                                 QMessageBox.Yes | QMessageBox.Cancel,
                                 QMessageBox.Cancel)  # "退出"代表的是弹出框的标题,"你确认退出.."表示弹出框的内容
        if a == QMessageBox.Yes:
            self.checkBox_ShowSegement.setEnabled(True)
            self.checkBox_ShowSegement.setChecked(True)
            for idx, imgarray in enumerate(self.image_array):
                image = cv2.resize(imgarray / 5, (512, 512), interpolation=cv2.INTER_NEAREST)
                cv2.imwrite(f'./temp/temp{idx}.png', image)
                img = Image.open(f'temp/temp{idx}.png')
                mask = self.predict_img(img)
                mask3 = mask[1].copy()
                self.segementmask.append(mask3)
            self.segementmask = np.array(self.segementmask)
            self.add_logs(f"All images({self.image_array.shape[0]}) have been segmented!")
        else:
            self.add_logs("You canceled the automatic segmentation.")
            pass

    #查看图片元数据
    def f_btn_showmetadata(self):
        self.MetaDataappw = MetaTableView()
        self.MetaDataappw.setWindowTitle('Dicom file MetaData')
        self.MetaDataappw.initData(self.metadata)
        self.MetaDataappw.show()

    #查看日志
    def f_btn_ViewLogs(self):
        self.ViewLogsWindow.textBrowser_logs.clear()
        for log in self.log_history:
            self.ViewLogsWindow.textBrowser_logs.append(log)
        self.ViewLogsWindow.show()

    #加载图片或图片序列

    def f_btn_loadimage(self):
        fname, _ = QFileDialog.getOpenFileName(self, 'Please select the medical image to be processed', current_directory,
                                               'Image files (*.dcm;*.IMA)')
        if fname:
            self.horizontalSlider_R_view.setEnabled(False)
            self.horizontalSlider_Y_view.setEnabled(False)
            self.horizontalSlider_Z_view.setEnabled(False)

            reader = sitk.ImageFileReader()

            reader.SetFileName(fname)

            reader.LoadPrivateTagsOn()

            reader.ReadImageInformation()
            rows = int(reader.GetMetaData("0028|0010"))  # Get number of rows from tag (0028, 0010)
            cols = int(reader.GetMetaData("0028|0011"))  # Get number of cols from tag (0028, 0011)
            for k in reader.GetMetaDataKeys():
                v = reader.GetMetaData(k)

                try:

                    self.metadata.append([f"({k})", v,
                                          self.tags[f"({k})".replace("|", ",").upper()][0],
                                          self.tags[f"({k})".replace("|", ",").upper()][1],
                                          self.tags[f"({k})".replace("|", ",").upper()][2],
                                          self.tags[f"({k})".replace("|", ",").upper()][3],
                                          self.tags[f"({k})".replace("|", ",").upper()][4],
                                          ])

                except:
                    self.metadata.append([f"({k})", v,
                                          "",
                                          "Not Found The DICOM Standard",
                                          "Non-standard",
                                          "",
                                          "",
                                          ])
            data = dicom_to_numpy(fname)
            mask = cv2.resize(data * 1.5, (512, 512), interpolation=cv2.INTER_NEAREST)
            cv2.imwrite('./temp3.png', mask)

            result = cv2.imread("./temp3.png", cv2.IMREAD_UNCHANGED)

            result = cv2.resize(result, (512, 512), interpolation=cv2.INTER_CUBIC)
            img_dis = QImage(result, result.shape[1], result.shape[0], result.shape[1],
                             QImage.Format_Grayscale8)
            img_dis = QPixmap(img_dis)
            self.view_R.setPixmap(img_dis)
            # print(rows,cols)

            self.add_logs(f'Load a image {fname},with shape{cols, rows}')

    def f_btn_loadimages(self):
        directory = QFileDialog.getExistingDirectory(None, "Please select a medical images folder path", current_directory)
        if directory:
            self.software_init()
            self.horizontalSlider_R_view.setEnabled(1)
            self.horizontalSlider_Y_view.setEnabled(1)
            self.horizontalSlider_Z_view.setEnabled(1)

            dirs, files = traversal_files(directory)

            for file in files:
                self.metadata.append(["", file.name, "", "", "", "", "", ])
                reader2 = sitk.ImageFileReader()
                reader2.SetFileName(file.path)
                reader2.LoadPrivateTagsOn()
                reader2.Execute()

                reader2.ReadImageInformation()

                for k in reader2.GetMetaDataKeys():
                    v = reader2.GetMetaData(k)
                    try:
                        self.metadata.append([f"({k})", v,
                                              self.tags[f"({k})".replace("|", ",").upper()][0],
                                              self.tags[f"({k})".replace("|", ",").upper()][1],
                                              self.tags[f"({k})".replace("|", ",").upper()][2],
                                              self.tags[f"({k})".replace("|", ",").upper()][3],
                                              self.tags[f"({k})".replace("|", ",").upper()][4],
                                              ])
                    except:
                        self.metadata.append([f"({k})", v, "", "Not Found The DICOM Standard", "Non-standard", "", "", ])

                if self.flag_Image_Position_readSuccessed:
                    try:
                        data0020032 = reader2.GetMetaData("0020|0032").split('\\')
                        self.Image_Position.append(
                            [float(data0020032[0]), float(data0020032[1]), float(data0020032[2]), ])  # Get number of rows from tag (0028, 0010)
                    except:
                        self.flag_Image_Position_readSuccessed = False
                        self.add_logs("Tag (0028|0032) missing!")
                        self.Image_Position.append(["Information Loss"])

                if self.flag_Image_Orientation_readSuccessed:
                    try:
                        data0020037 = reader2.GetMetaData("0020|0037").split('\\')
                        self.Image_Orientation.append(
                            [float(data0020037[0]), float(data0020037[1]), float(data0020037[2]), float(data0020037[3]), float(data0020037[4]),
                             float(data0020037[5]), ])  # Get number of cols from tag (0028, 0011)
                    except:
                        self.flag_Image_Orientation_readSuccessed = False
                        self.add_logs("Tag (0028|0037) missing!")
                        self.Image_Orientation.append(["Information Loss"])

                if self.flag_Pixel_Spacing_readSuccessed:
                    try:
                        data0020030 = reader2.GetMetaData("0028|0030").split('\\')
                        self.Pixel_Spacing.append([float(data0020030[0]), float(data0020030[1])])
                    except:
                        self.flag_Pixel_Spacing_readSuccessed = False
                        self.add_logs("Tag (0028|0030) missing!")
                        self.Pixel_Spacing.append(["Information Loss"])

                if self.flag_Patient_Position_readSuccessed:
                    try:
                        data00185100 = reader2.GetMetaData("0018|5100").split('\\')
                        self.Patient_Position.append(data00185100)
                    except:
                        self.flag_Patient_Position_readSuccessed = False
                        self.add_logs("Tag (0028|5100) missing!")
                        self.Patient_Position.append(["Information Loss"])

            spacing = [0.75, 0.75, 1]
            for position in self.Image_Position:
                origin = position
                lt_x_real = origin[0]
                lt_y_real = origin[1] + 0 * spacing[1]
                lt_z_real = origin[2] + 0 * spacing[2]

                rt_x_real = origin[0]
                rt_y_real = origin[1] + 191 * spacing[1]
                rt_z_real = origin[2] + 0 * spacing[2]

                rb_x_real = origin[0]
                rb_y_real = origin[1] + 191 * spacing[1]
                rb_z_real = origin[2] + 191 * spacing[2]

                lb_x_real = origin[0]
                lb_y_real = origin[1] + 0 * spacing[1]
                lb_z_real = origin[2] + 191 * spacing[2]
                self.Rcorners.append([[lt_x_real, lt_y_real, lt_z_real], [rt_x_real, rt_y_real, rt_z_real]
                                         , [rb_x_real, rb_y_real, rb_z_real], [lb_x_real, lb_y_real, lb_z_real]])

            tp1, tp2, bp2, bp1 = self.Rcorners[0]
            tp4, tp3, bp3, bp4 = self.Rcorners[-1]

            tttz = (bp4[2] - tp4[2]) / 191.0
            ttty = (tp3[1] - tp4[1]) / 191.0

            # print(tp1[2] - tp4[2], tp2[1] - tp1[1], ttty, tttz)
            for i in range(192):
                self.Zcorners.append([[tp4[0], tp4[1], tp4[2] + tttz * i], [tp3[0], tp3[1], tp3[2] + tttz * i],
                                      [tp2[0], tp2[1], tp2[2] + tttz * i], [tp1[0], tp1[1], tp1[2] + tttz * i]])
                self.Ycorners.append([[tp4[0], tp4[1] + 1 + ttty * i, tp4[2]], [tp1[0], tp1[1] + ttty * i, tp1[2], ],
                                      [bp1[0], bp1[1] + 1 + ttty * i, bp1[2]], [bp4[0], bp4[1] + ttty * i, bp4[2], ]])

            self.label_ImagePositonFirst.setText(
                f"{round(self.Image_Position[0][0], 1), round(self.Image_Position[0][1], 1), round(self.Image_Position[0][2], 1),}")
            self.label_ImagePositonLast.setText(
                f"{round(self.Image_Position[-1][0], 1), round(self.Image_Position[-1][1], 1), round(self.Image_Position[-1][2], 1),}")

            reader = sitk.ImageSeriesReader()
            img_names = reader.GetGDCMSeriesFileNames(directory)
            reader.SetFileNames(img_names)
            image = reader.Execute()

            # 获取图像序列
            self.image_array = sitk.GetArrayFromImage(image)

            # 设定各滑动条最大值
            self.horizontalSlider_R_view.setMaximum(self.image_array.shape[0] - 1)  # z, y, x
            self.horizontalSlider_Y_view.setMaximum(self.image_array.shape[2] - 1)  # z, y, x
            self.horizontalSlider_Z_view.setMaximum(self.image_array.shape[1] - 1)  # z, y, x

            self.f_horizontalSlider_R_view(0)
            self.f_horizontalSlider_Y_view(0)
            self.f_horizontalSlider_Z_view(0)

            self.add_logs(f'Load a Serial image data {directory},with shape{self.image_array.shape}')
        self.f_btn_reset3dview()

    #标记点

    def f_btn_MarkMarker1(self):
        self.markernumber = 1
        self.markmarker()

    def f_btn_MarkMarker2(self):
        self.markernumber = 2
        self.markmarker()

    def f_btn_MarkMarker3(self):
        self.markernumber = 3
        self.markmarker()

    def f_btn_MarkMarker4(self):
        self.markernumber = 4
        self.markmarker()

    def f_btn_MarkProstateCenter(self):
        self.markernumber = 5
        self.markmarker()

    def eventFilter(self, obj, event):
        if obj == self.view_R and event.type() == QtCore.QEvent.MouseButtonPress:
            # 点击事件发生时，获取点击的坐标
            x = event.pos().x()
            y = event.pos().y()

            View_R_w, View_R_h = self.view_R.width(), self.view_R.height()
            pic_w, pic_h = self.image_array.shape[1], self.image_array.shape[2]

            if View_R_h < View_R_w:
                xdiff = (View_R_w - View_R_h) / 2
                x = x - xdiff
            if View_R_w < View_R_h:
                ydiff = (View_R_h - View_R_w) / 2
                y = y - ydiff
            whminal = min(View_R_w, View_R_h)

            x = x * pic_w / whminal
            y = y * pic_h / whminal
            spacing = [0.75, 0.75, 1]
            origin = self.Image_Position[self.current_R_view_frame]
            lt_x_real = origin[0]
            lt_y_real = origin[1] + x * spacing[1]
            lt_z_real = origin[2] + y * spacing[2]
            # print(xyz)

            if self.markernumber != 5:
                self.add_logs(f"Marker{self.markernumber}:You set a marker at {lt_x_real, lt_y_real, lt_z_real}mm position.")
                self.handlemarkercreate((lt_x_real, lt_y_real, lt_z_real))
            if self.markernumber == 5:
                self.handleMarkProstateCenterCreate((lt_x_real, lt_y_real, lt_z_real))
            self.view_R.removeEventFilter(self)  # 安装事件过滤器
            QApplication.restoreOverrideCursor()
            return True  # 事件已被处理，不再传递
        return super().eventFilter(obj, event)

    def markmarker(self):
        if self.image_array.any() != None:
            self.view_R.removeEventFilter(self)
            self.view_R.installEventFilter(self)
            QApplication.setOverrideCursor(self.cursor)
        else:
            self.add_logs("Haven't loaded the picture yet, please load the picture and then mark the Marker point.")

    def handlemarkercreate(self, markerposition):

        if self.markernumber == 1:
            try:
                self.ren.RemoveActor(self.marker1)
            except:
                pass
            self.marker1 = DrawAMarker(point=(markerposition[0], markerposition[1], markerposition[2]), Radius=5)
            self.marker_position[f'mark1'] = [markerposition[0], markerposition[1], markerposition[2]]
            self.label_marker1info.setText(f"{round(markerposition[0], 2), round(markerposition[1], 2), round(markerposition[2], 2)}")
            self.ren.AddActor(self.marker1)
            self.checkMarker()
            self.vtkWidget.GetRenderWindow().Render()
        if self.markernumber == 2:
            try:
                self.ren.RemoveActor(self.marker2)
            except:
                pass
            self.marker2 = DrawAMarker(point=(markerposition[0], markerposition[1], markerposition[2]), Radius=5)
            self.marker_position[f'mark2'] = [markerposition[0], markerposition[1], markerposition[2]]
            self.label_marker2info.setText(f"{round(markerposition[0], 2), round(markerposition[1], 2), round(markerposition[2], 2)}")
            self.ren.AddActor(self.marker2)
            self.checkMarker()
            self.vtkWidget.GetRenderWindow().Render()
        if self.markernumber == 3:
            try:
                self.ren.RemoveActor(self.marker3)
            except:
                pass
            self.marker3 = DrawAMarker(point=(markerposition[0], markerposition[1], markerposition[2]), Radius=5)
            self.marker_position[f'mark3'] = [markerposition[0], markerposition[1], markerposition[2]]
            self.label_marker3info.setText(f"{round(markerposition[0], 2), round(markerposition[1], 2), round(markerposition[2], 2)}")
            self.ren.AddActor(self.marker3)
            self.checkMarker()
            self.vtkWidget.GetRenderWindow().Render()
        if self.markernumber == 4:
            try:
                self.ren.RemoveActor(self.marker4)
            except:
                pass
            self.marker4 = DrawAMarker(point=(markerposition[0], markerposition[1], markerposition[2]), Radius=5)
            self.marker_position[f'mark4'] = [markerposition[0], markerposition[1], markerposition[2]]
            self.label_marker4info.setText(f"{round(markerposition[0], 2), round(markerposition[1], 2), round(markerposition[2], 2)}")
            self.ren.AddActor(self.marker4)
            self.checkMarker()
            self.vtkWidget.GetRenderWindow().Render()

    def handleMarkProstateCenterCreate(self, markerposition):
        try:
            self.ren.RemoveActor(self.ProstateCenter)
        except:
            pass
        self.ProstateCenter = DrawProstateCenter(point=(markerposition[0], markerposition[1], markerposition[2]))
        self.label_ProstatePosition.setText(f"{round(markerposition[0], 2), round(markerposition[1], 2), round(markerposition[2], 2)}")
        self.add_logs(f"You set Prostate Center at {markerposition[0], markerposition[1], markerposition[2]}mm position.")
        self.ren.AddActor(self.ProstateCenter)
        self.vtkWidget.GetRenderWindow().Render()

    def checkMarker(self):
        error = MarkercheckAixs(self.marker_position[f'mark1'], self.marker_position[f'mark2'],
                                self.marker_position['mark3'], self.marker_position['mark4'])
        self.label_markcheckerror.setText(f"{error:.4f}")

    #RYZ绘制切片角点和平面

    def Drawplanep1p2p3p4x(self, p1, p2, p3, p4):
        try:
            self.ren.RemoveActor(self.XPlaneActor)
        except:
            pass
        self.XPlaneActor = DrawaPlanep1p2p3p4(p1=p1, p2=p2, p3=p3, p4=p4, corlors='red')

        self.ren.AddActor(self.XPlaneActor)
        self.vtkWidget.GetRenderWindow().Render()

    def DrawpPointp1p2p3p4x(self, p1, p2, p3, p4):
        try:
            self.ren.RemoveActor(self.XPlaneActorp1)
            self.ren.RemoveActor(self.XPlaneActorp2)
            self.ren.RemoveActor(self.XPlaneActorp3)
            self.ren.RemoveActor(self.XPlaneActorp4)
        except:
            pass
        self.XPlaneActorp1 = DrawAMarker(p1, color='red')
        self.XPlaneActorp2 = DrawAMarker(p2, color='red')
        self.XPlaneActorp3 = DrawAMarker(p3, color='red')
        self.XPlaneActorp4 = DrawAMarker(p4, color='red')

        self.ren.AddActor(self.XPlaneActorp1)
        self.ren.AddActor(self.XPlaneActorp2)
        self.ren.AddActor(self.XPlaneActorp3)
        self.ren.AddActor(self.XPlaneActorp4)
        self.vtkWidget.GetRenderWindow().Render()

    def Drawplanep1p2p3p4y(self, p1, p2, p3, p4):
        try:
            self.ren.RemoveActor(self.YPlaneActor)
        except:
            pass
        self.YPlaneActor = DrawaPlanep1p2p3p4(p1=p1, p2=p2, p3=p3, p4=p4, corlors='green')

        self.ren.AddActor(self.YPlaneActor)
        self.vtkWidget.GetRenderWindow().Render()

    def DrawpPointp1p2p3p4y(self, p1, p2, p3, p4):
        try:
            self.ren.RemoveActor(self.YPlaneActorp1)
            self.ren.RemoveActor(self.YPlaneActorp2)
            self.ren.RemoveActor(self.YPlaneActorp3)
            self.ren.RemoveActor(self.YPlaneActorp4)
        except:
            pass
        self.YPlaneActorp1 = DrawAMarker(p1, color='green')
        self.YPlaneActorp2 = DrawAMarker(p2, color='green')
        self.YPlaneActorp3 = DrawAMarker(p3, color='green')
        self.YPlaneActorp4 = DrawAMarker(p4, color='green')

        self.ren.AddActor(self.YPlaneActorp1)
        self.ren.AddActor(self.YPlaneActorp2)
        self.ren.AddActor(self.YPlaneActorp3)
        self.ren.AddActor(self.YPlaneActorp4)
        self.vtkWidget.GetRenderWindow().Render()

    def Drawplanep1p2p3p4z(self, p1, p2, p3, p4):
        try:
            self.ren.RemoveActor(self.ZPlaneActor)
        except:
            pass
        self.ZPlaneActor = DrawaPlanep1p2p3p4(p1=p1, p2=p2, p3=p3, p4=p4, corlors='blue')

        self.ren.AddActor(self.ZPlaneActor)
        self.vtkWidget.GetRenderWindow().Render()

    def DrawpPointp1p2p3p4z(self, p1, p2, p3, p4):
        try:
            self.ren.RemoveActor(self.ZPlaneActorp1)
            self.ren.RemoveActor(self.ZPlaneActorp2)
            self.ren.RemoveActor(self.ZPlaneActorp3)
            self.ren.RemoveActor(self.ZPlaneActorp4)
        except:
            pass
        self.ZPlaneActorp1 = DrawAMarker(p1, color='blue')
        self.ZPlaneActorp2 = DrawAMarker(p2, color='blue')
        self.ZPlaneActorp3 = DrawAMarker(p3, color='blue')
        self.ZPlaneActorp4 = DrawAMarker(p4, color='blue')

        self.ren.AddActor(self.ZPlaneActorp1)
        self.ren.AddActor(self.ZPlaneActorp2)
        self.ren.AddActor(self.ZPlaneActorp3)
        self.ren.AddActor(self.ZPlaneActorp4)
        self.vtkWidget.GetRenderWindow().Render()

    # RYZ窗口切换视图

    def f_horizontalSlider_R_view(self, sigint):

        self.label_frames_R_view.setText(f'{sigint + 1}/{self.image_array.shape[0]}')
        self.label_depth_R_view.setText(f"X(R->L):{self.Image_Position[sigint][0]:.4f}mm")
        self.current_slice_position_x = self.Image_Position[sigint]
        self.current_R_view_frame = sigint

        p1, p2, p3, p4 = self.Rcorners[sigint]
        self.DrawpPointp1p2p3p4x(p1=p1, p2=p2, p3=p3, p4=p4)
        self.Drawplanep1p2p3p4x(p1=p1, p2=p2, p3=p3, p4=p4)

        self.label_RView_LeftTop.setText(f"{round(p1[0], 2), round(p1[1], 2), round(p1[2], 2)}")
        self.label_RView_RightTop.setText(f"{round(p2[0], 2), round(p2[1], 2), round(p2[2], 2)}")
        self.label_RView_RightBottom.setText(f"{round(p3[0], 2), round(p3[1], 2), round(p3[2], 2)}")
        self.label_RView_LeftBottom.setText(f"{round(p4[0], 2), round(p4[1], 2), round(p4[2], 2)}")

        self.label_SliceInfoImagePosition_x.setText(
            f"{round(self.Image_Orientation[sigint][0], 2), round(self.Image_Orientation[sigint][1], 2), round(self.Image_Orientation[sigint][2], 2)}")
        self.label_SliceInfoImagePosition_y.setText(
            f"{round(self.Image_Orientation[sigint][3], 2), round(self.Image_Orientation[sigint][4], 2), round(self.Image_Orientation[sigint][5], 2)}")
        self.label_SliceInfoPixelSpacing_xy.setText(f"{round(self.Pixel_Spacing[sigint][0], 4), round(self.Pixel_Spacing[sigint][1], 4)}")
        #self.label_SliceInfoPatientPosition.setText(self.Patient_Position[0])

        Rwidth, Rheight = self.view_R.width(), self.view_R.height()
        picw, pich = self.image_array.shape[1], self.image_array.shape[2],
        xishu = min(Rwidth / picw, Rheight / pich)
        image = cv2.resize(self.image_array[sigint] / 5, (int(pich * xishu), int(picw * xishu)), interpolation=cv2.INTER_NEAREST)
        cv2.imwrite('./temp/Rimage.png', image)
        result = cv2.imread("./temp/Rimage.png", cv2.IMREAD_UNCHANGED)
        src_RGB = cv2.cvtColor(result, cv2.COLOR_GRAY2BGR)

        if self.checkBox_ShowSegement.isChecked():
            src_RGB = put_mask(src_RGB, self.segementmask[sigint])
        img_dis = QImage(src_RGB, src_RGB.shape[1], src_RGB.shape[0], src_RGB.shape[1] * 3, QImage.Format_RGB888)
        img_dis = QPixmap(img_dis).scaled(src_RGB.shape[1], src_RGB.shape[0])
        self.view_R.setPixmap(img_dis)

    def f_horizontalSlider_Y_view(self, sigint):
        p1, p2, p3, p4 = self.Ycorners[sigint]
        self.DrawpPointp1p2p3p4y(p1=p1, p2=p2, p3=p3, p4=p4)
        self.Drawplanep1p2p3p4y(p1=p1, p2=p2, p3=p3, p4=p4)

        self.label_YView_LeftTop.setText(f"{round(p1[0], 2), round(p1[1], 2), round(p1[2], 2)}")
        self.label_YView_RightTop.setText(f"{round(p2[0], 2), round(p2[1], 2), round(p2[2], 2)}")
        self.label_YView_RightBottom.setText(f"{round(p3[0], 2), round(p3[1], 2), round(p3[2], 2)}")
        self.label_YView_LeftBottom.setText(f"{round(p4[0], 2), round(p4[1], 2), round(p4[2], 2)}")

        self.label_frames_Y_view.setText(f'{sigint}/{self.image_array.shape[1]}')

        # self.label_depth_Y_view.setText(f"Y(A->P):{self.Image_Position[sigint][1]:.4f}mm")
        Gimage = self.image_array[:, sigint, :] / 5

        Gwidth, Gheight = self.view_Y.width(), self.view_Y.height()
        picw, pich = Gimage.shape[0], Gimage.shape[1]
        xishu = min(Gwidth / picw, Gheight / pich)
        Gimage = cv2.resize(Gimage, (int(pich * xishu), int(picw * xishu)))

        cv2.imwrite('./temp/Gimage.png', Gimage)
        Gimage = cv2.imread("./temp/Gimage.png", cv2.IMREAD_UNCHANGED)
        src_RGB = cv2.cvtColor(Gimage, cv2.COLOR_GRAY2BGR)
        if self.checkBox_ShowSegement.isChecked():
            src_RGB = put_mask(src_RGB, self.segementmask[:, sigint, :])

        img_dis = QImage(src_RGB, src_RGB.shape[1], src_RGB.shape[0], src_RGB.shape[1] * 3,
                         QImage.Format_RGB888)
        img_dis = QPixmap(img_dis).scaled(src_RGB.shape[1], src_RGB.shape[0])
        self.view_Y.setPixmap(img_dis)

    def f_horizontalSlider_Z_view(self, sigint):
        "Z(I->S)"
        p1, p2, p3, p4 = self.Zcorners[sigint]
        self.DrawpPointp1p2p3p4z(p1=p1, p2=p2, p3=p3, p4=p4)
        self.Drawplanep1p2p3p4z(p1=p1, p2=p2, p3=p3, p4=p4)

        self.label_ZView_LeftTop.setText(f"{round(p1[0], 2), round(p1[1], 2), round(p1[2], 2)}")
        self.label_ZView_RightTop.setText(f"{round(p2[0], 2), round(p2[1], 2), round(p2[2], 2)}")
        self.label_ZView_RightBottom.setText(f"{round(p3[0], 2), round(p3[1], 2), round(p3[2], 2)}")
        self.label_ZView_LeftBottom.setText(f"{round(p4[0], 2), round(p4[1], 2), round(p4[2], 2)}")

        self.label_frames_Z_view.setText(f'{sigint}/{self.image_array.shape[2]}')
        self.label_depth_Z_view.setText(
            f"Z(I->S):{self.Image_Position[-1][2] + (self.Image_Position[-1][2] - self.Image_Position[0][2]) / self.image_array.shape[2]:.4f}mm")
        Gimage = self.image_array[:, :, sigint] / 5

        Gwidth, Gheight = self.view_Z.width(), self.view_Z.height()
        picw, pich = Gimage.shape[0], Gimage.shape[1]
        xishu = min(Gwidth / picw, Gheight / pich)
        Gimage = cv2.resize(Gimage, (int(pich * xishu), int(picw * xishu)))

        cv2.imwrite('./temp/Zimage.png', Gimage)
        Gimage = cv2.imread("./temp/Zimage.png", cv2.IMREAD_UNCHANGED)
        src_RGB = cv2.cvtColor(Gimage, cv2.COLOR_GRAY2BGR)
        if self.checkBox_ShowSegement.isChecked():
            src_RGB = put_mask(src_RGB, self.segementmask[:, :, sigint])

        img_dis = QImage(src_RGB, src_RGB.shape[1], src_RGB.shape[0], src_RGB.shape[1] * 3,
                         QImage.Format_RGB888)
        img_dis = QPixmap(img_dis).scaled(src_RGB.shape[1], src_RGB.shape[0])
        self.view_Z.setPixmap(img_dis)

    # FullScreen View

    def f_btn_FullScreen_R_view(self):
        self.view_R_fullScreen.show()
        self.add_logs('Sorry, the full -screen display view function has not been completed.')

    def f_btn_FullScreen_Y_view(self):
        self.view_Y_fullScreen.show()
        self.add_logs('Sorry, the full -screen display view function has not been completed.')

    def f_btn_FullScreen_Z_view(self):
        self.view_Z_fullScreen.show()
        self.add_logs('Sorry, the full -screen display view function has not been completed.')

    def f_btn_FullScreen_3D_view(self):
        pass


app = QApplication(sys.argv)
w = MyApp()
w.show()
sys.exit(app.exec())
