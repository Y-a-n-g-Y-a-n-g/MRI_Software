import pyautogui
import requests
import win32gui
from PIL import ImageGrab
from fuzzywuzzy import fuzz
from openpyxl import load_workbook
import cv2 as cv
import numpy as np
import aircv as ac
import time
import string
from zhon.hanzi import punctuation
import threading
import sys
import os
from PIL import Image
from PIL import ImageChops
from pynput.mouse import Button, Controller as c_mouse
import win32com.client
#测试标志 1 测试状态  0 正式答题
testflag=0
#设置自建服务器
#cpu python Tr/backend/main.py
#gpu python Tr_GPU/backend/main.py;nohup python3 ocr/backend/main.py
zijianURL=['http://43.138.168.143:8089/api/tr-run/','http://116.85.57.11:8089/api/tr-run/','http://116.85.7.193:8089/api/tr-run/','http://116.85.42.68:8089/api/tr-run/','http://116.85.68.165:8089/api/tr-run/']
#设置答题服务器数量
zijianURL_numbers=1
# pyinstaller -i ico.ico -D --add-data '.\最新五合一题库.xlsx;.' --add-data '.\data;.' .\答题.py
#设置答题抓取窗口
if testflag==1:
    titlename='梦幻西游自动答题训练器'
else:
    titlename = '梦幻西游 online'


#设置答题延时
print("答题延时,1=0.01，2=0.02，3=0.03，4=0.04，5=0.05,(默认1)")
time_dati_delay_check=input('答题延时=')
time_dati_delay=0.01
if time_dati_delay_check=='1':
    time_dati_delay=0.01
elif time_dati_delay_check=='2':
    time_dati_delay=0.02
elif time_dati_delay_check=='3':
    time_dati_delay=0.03
elif time_dati_delay_check=='4':
    time_dati_delay=0.04
elif time_dati_delay_check=='5':
    time_dati_delay=0.05
#设置点击延时
print("点击延时,1=0.05，2=0.02，3=0.03，4=0.04，5=0.08,6=0.1,7=0.15 (默认1)")
time_click_delay_check=input('答题延时=')
time_click_delay=0.05
if time_click_delay_check=='1':
    time_click_delay=0.05
elif time_click_delay_check=='2':
    time_click_delay=0.02
elif time_click_delay_check=='3':
    time_click_delay=0.03
elif time_click_delay_check=='4':
    time_click_delay=0.04
elif time_click_delay_check=='5':
    time_click_delay=0.08
elif time_click_delay_check=='6':
    time_click_delay=0.1
elif time_click_delay_check=='7':
    time_click_delay=0.15
#获取答题题库 1=殿试，2=乡试，3=房都尉，4=生活常识，5=元宵节答题
tikunames=['1=殿试所有题','2=乡试所有题','3=房都尉题库','4=生活常识-附加题','5=元宵节所有题']
print("答题题库,1=殿试，2=乡试，3=房都尉，4=生活常识，5=元宵节答题")
tiku=input('题库=')
while tiku not in ['1','2','3','4','5']:
    print("答题题库,1=殿试，2=乡试，3=房都尉，4=生活常识，5=元宵节答题")
    tiku = input('题库=')

tiku=tikunames[int(tiku)-1]

#已选择窗口
xuanze_titlename=None
xuanze_hwnd=None
windows = None
mouse= c_mouse()

def move( x, y):

    # pyautogui.moveTo(x, y)
    mouse.position = (x, y)

def click(x, y):
    move(x, y)
    time.sleep(0.025)
   # pyautogui.click(clicks=2)
    ##按下右键
    mouse.press(Button.left)
    ##释放右键
    mouse.release(Button.left)
    time.sleep(0.025)
if getattr(sys, 'frozen', None):
    basedir = sys._MEIPASS
else:
    basedir = os.path.dirname(__file__)





#清除标点符号
def clearpunctuation(str=''):
    for i in string.punctuation:
        str = str.replace(i, '')
    for i in punctuation:
        str = str.replace(i, '')
    str = str.replace(' ', '')
    return str
#给出答题链接
class GetURL():
    def __init__(self):
        self.zijianURL=zijianURL
        self.zijianURLflag=0
    def url(self):
        self.zijianURLflag+=1
        self.zijianURLflag=0 if self.zijianURLflag==zijianURL_numbers else self.zijianURLflag
        return self.zijianURL[self.zijianURLflag]
geturl=GetURL()
class Pic2txt():
    def pic2txt(self,img):
        '''
        自建服务
        传入:numpy格式的图片
        传出:成功:服务器返回的josn数据
        '''
        res = requests.post(url=geturl.url(), data={'compress': 0}, files=img)
        return self.printwords(res.json()['data']['raw_out'])
    def printwords(self,items):
        '''
        自建服务
        传入:自建服务中返回的数据
        传出:成功:服务器返回的题目
        '''
        result = ''
        for item in items:
            result = result + item[1]
        if '题:' in result:
            result = result.split("题:")[-1]
        if '目:' in result:
            result = result.split("目:")[-1]
        if '题。' in result:
            result = result.split("题。")[-1]
        return result

class Get_Screen():
    """截图程序"""
    def __init__(self,):
        """
        窗口初始化函数
        传入参数:default_titlename 窗口名
        成功:self.initflag=1
        失败:self.initflag = 0
        """
        global xuanze_titlename,xuanze_hwnd,windows
        self.toplist, self.winlist = [], []
        win32gui.EnumWindows(self.enum_cb, self.toplist)
        menhuanxiyou = [(hwnd, title) for hwnd, title in self.winlist if  titlename in title.lower()]
        if len(menhuanxiyou)>1 or not xuanze_hwnd:
            if not xuanze_hwnd:
                print("检测到多个残留进程,请选择正确的进程:")
                for idx,(hwnd,title) in enumerate(menhuanxiyou):
                    print(f"{idx}. {title}")
                idx=int(input("请输入:\n"))
                self.title=menhuanxiyou[idx][1]
                self.hwnd = menhuanxiyou[idx][0]
                windows = ( self.hwnd,self.title)
                xuanze_titlename=self.title
                xuanze_hwnd=self.hwnd
        self.initflag=1
    def enum_cb(self,hwnd, results):
        self.winlist.append((hwnd, win32gui.GetWindowText(hwnd)))
    def get_bbox(self):
        """
        返回窗口的绝对位置
        传出:
            成功:返回窗口坐标
            失败:返回0
        """
        self.__init__()
        try:
            if self.initflag:
                return win32gui.GetWindowRect(self.hwnd)
        except:
            return 0
    def get_screen(self,bbox=None):
        """
        截取屏幕某一区域图像
        传入:[屏幕左上坐标,屏幕右下坐标] 传入空参默认截取初始化函数中的区域图像
        传出:
            成功:返回[区域图像,窗口的坐标]
            失败:返回0
        """
        try:
            if bbox:
                return [ImageGrab.grab(bbox), bbox]
            else:
                return [ImageGrab.grab(self.get_bbox()), bbox]
        except:
            return 0
class Findkey():
    '''匹配答案-多答案版本'''

    def __init__(self):
        """
        初始化答案匹配类
        传入: filename题库文件名 sheet_ranges:题库表名
        成功:self.initflag=1
        失败:self.initflag=0
        """
        # try:
        tikuwenjian = os.path.join(basedir, '最新五合一题库.xlsx')
        wb = load_workbook(tikuwenjian)
        sheet_ranges = wb[tiku]
        self.questions = {}
        self.abcd = ['A', 'B', 'C', 'D']
        for i in range(1, sheet_ranges.max_row + 1):
            self.questions[clearpunctuation(str(sheet_ranges['A' + str(i)].value))] = [
                clearpunctuation((str(sheet_ranges['B' + str(i)].value))), ]
            try:
                lie = 'C'
                while sheet_ranges[lie + str(i)].value:
                    list = self.questions[clearpunctuation(str(sheet_ranges['A' + str(i)].value))]
                    list.append(clearpunctuation((str(sheet_ranges[lie + str(i)].value))))
                    self.questions[clearpunctuation(str(sheet_ranges['A' + str(i)].value))] = list
                    lie = str(chr(ord(lie) + 1))
            except:
                pass

        self.questionsonly = [clearpunctuation(data) for data in self.questions.keys()]
        self.answeeronly = []
        sheet_ranges = wb['6-殿试单独答案']
        for i in range(1, sheet_ranges.max_row + 1):
            self.answeeronly.append(clearpunctuation((str(sheet_ranges['A' + str(i)].value))))
        #print(self.answeeronly)
    def getanswerfromquestion(self, question='', keyA='', keyB='', keyC='', keyD=''):
        """
        匹配答案
        传入:图片识别得到的问题和四个选项
        传出:成功:识别到的选项,匹配的问题和答案
            异常:返回0
        """
        # try:
        values1 = [fuzz.ratio(question, data) for data in self.questionsonly]
        question = self.questionsonly[values1.index(max(values1))]
        answers = self.questions[question]
        maxvalue = 0
        maxvalueanswers = answers[0]
        keys = [keyA, keyB, keyC, keyD]
        for answer in answers:
            values2 = [fuzz.ratio(answer, data) for data in keys]
            maxvalueanswers = answer if max(values2) > maxvalue else maxvalueanswers
            maxvalue = max(values2) if max(values2) > maxvalue else maxvalue
        values2 = [fuzz.ratio(maxvalueanswers, data) for data in keys]
        return [self.abcd[values2.index(max(values2))], question, maxvalueanswers, max(values1), max(values2)]

    def getanswerfromkey(self, keyA='', keyB='', keyC='', keyD=''):
        """
        匹配答案
        传入:图片识别得到的问题和四个选项
        传出:成功:识别到的选项,匹配的问题和答案
            异常:返回0
        """
        try:
            if max([fuzz.ratio(keyA, data) for data in self.answeeronly]) == 100:
                return ['A', '', keyA, 100]
            if max([fuzz.ratio(keyB, data) for data in self.answeeronly]) == 100:
                return ['B', '', keyB, 100]
            if max([fuzz.ratio(keyC, data) for data in self.answeeronly]) == 100:
                return ['C', '', keyC, 100]
            if max([fuzz.ratio(keyD, data) for data in self.answeeronly]) == 100:
                return ['D', '', keyD, 100]
            return 0
        except:
            return 0
class answerAPI():
    def __init__(self):
        self.tiku = tiku
        if self.tiku == '3=房都尉题库':
            self.startwen = cv.imread('data/fangduwei.png')
            self.leave = cv.imread('data/fangduweileave.png')
        elif self.tiku == '4=生活常识-附加题':
            self.startwen = cv.imread('data/fujiati.png')
            self.leave = cv.imread('data/fujiatileave.png')
        elif self.tiku == '1=殿试所有题':
            self.startwen = cv.imread('data/dianshi.png')
            self.startwen37 = cv.imread('data/dianshi37.png')
            self.leave = cv.imread('data/dianshileave.png')
        elif self.tiku == '2=乡试所有题':
            pass#待写
        elif self.tiku == '5=元宵节所有题':
            pass#待写

    # 通用API
    def get_areaready(self, mubiao, moban, confidence=0.7):
        """
        通用返回特定图像中是否存在某图像
        传入:待检测图像,需检测图像,自信度
        传出:存在:返回1
            不存在或错误:返回0
        """
        try:
            match_result = ac.find_template(mubiao, moban, confidence)
            #print(match_result)
            if match_result is not None:
                match_result['shape'] = (mubiao.shape[1], moban.shape[0])
                return 1
            else:
                return 0
        except:
            return 0
    def get_diff(self, pictrue1,pictrue2):
        """
        通用返回两图片是否相同
        传入:待检测图像1,待检测图像2
        传出:相同:返回1
            不存在或错误:返回0
        """
        try:
            im1 = Image.open(pictrue1)
            im2 = Image.open(pictrue2)
            diff = ImageChops.difference(im2, im1).getbbox()
            #print(diff)
            if diff:
                return 0
            else:
                return 1
        except:
            return 0

    def gettxtfromAPI_thread(self,img,num):
        img = {'file': open(f'./data/{img}.png', 'rb')}
        result=clearpunctuation(pic2txt.pic2txt(img))
        if num==0:
            self.txt_wens=result
        elif num==1:
            self.txt_As = result
        elif num==2:
            self.txt_Bs = result
        elif num==3:
            self.txt_Cs = result
        elif num==4:
            self.txt_Ds = result
    def gettxtfromAPI(self):
        '''
        调用自建APi识别题目和选项
        传入:wen,A,B,C,D的图像
        传出:成功:wen,A,B,C,D的文本
            异常:返回0
        '''
        #try:
        tlist = []
        cv.imwrite('./data/wen.png', self.handleed_wen)
        cv.imwrite('./data/A.png', self.handleed_A)
        cv.imwrite('./data/B.png', self.handleed_B)
        cv.imwrite('./data/C.png', self.handleed_C)
        cv.imwrite('./data/D.png',self.handleed_D)
        self.txt_wens, self.txt_As, self.txt_Bs, self.txt_Cs, self.txt_Ds='','','','',''
        for img,result in zip(['wen','A','B','C','D'],[0,1,2,3,4]):
            t = threading.Thread(target=self.gettxtfromAPI_thread, args=(img,result))  # 创建线程
            t.start()
            tlist.append(t)
        for t in tlist:
            t.join()
        return 1



    # 自动判断调用
    def jiancekaishi(self):
        if self.tiku == '3=房都尉题库':
            return self.jiancekaishi_fangduwei()
        elif self.tiku == '1=殿试所有题':
            return self.jiancekaishi_dianshi()
        elif self.tiku == '4=生活常识-附加题':
            return self.jiancekaishi_shenghuofujiati()
        else:
            return 0
    def jiancekaishi2(self):
        try:
            if self.tiku == '3=房都尉题库':
                return self.jiancekaishi_fangduwei2()
            elif self.tiku == '1=殿试所有题':
                return self.jiancekaishi_dianshi2()
            elif self.tiku == '4=生活常识-附加题':
                return self.jiancekaishi_shenghuofujiati2()
            else:
                return 0
        except:
            return 0

    def huida(self, xuangxiang):
        try:
            if self.tiku == '3=房都尉题库':
                return self.huida_fangduwei(xuangxiang)
            elif self.tiku == '1=殿试所有题':
                return self.huida_dianshi(xuangxiang)
        except:
            return 0

    def hanleimgquestion(self):
        try:
            if self.tiku == '3=房都尉题库':
                return self.handleimgquestion_fangduwei()
            elif self.tiku == '1=殿试所有题':
                return self.handleimgquestion_dianshi()
            elif self.tiku == '4=生活常识-附加题':
                return self.handleimgquestion_shenghuochangshi()
            else:
                return 0
        except:
            return 0

    # 针对殿试
    def huida_dianshi(self, xuangxiang):
        """
        自动回答程序
        传入:选项
        传出:成功 1
            失败 0
        """
        try:
            x, y, *_ = get_screen.get_bbox()
            if self.flag37==0:
                if xuangxiang == 'A':
                    click(x + 307, y + 319)
                elif xuangxiang == 'B':
                    click(x + 520, y + 320)
                elif xuangxiang == 'C':
                    click(x + 307, y + 395)
                elif xuangxiang == 'D':
                    click(x + 520, y + 395)
            else:
                if xuangxiang == 'A':
                    click(x + 394, y + 355)
                elif xuangxiang == 'B':
                    click(x + 538, y + 351)
                elif xuangxiang == 'C':
                    click(x + 397, y + 429)
                elif xuangxiang == 'D':
                    click(x + 538, y + 426)
            time.sleep(time_click_delay)
            move(x + 130, y + 320)
            time.sleep(0.01)
            return 1
        except:
            return 0

    def handleimgquestion_dianshi(self):
        """
        针对房都蔚划分区域
        输入:numpy格式的图像
        输出:二值化后的问题,选项ABCD的图像
        """
        try:

            mubiao, _ = get_screen.get_screen()
            mubiao = cv.cvtColor(np.asarray(mubiao), cv.COLOR_RGB2BGR)
            if self.flag37==0:
                wen = [223+15, 110 + 60, 620, 190 + 60]
                A = [223+15, 235 + 60, 400, 285 + 60]
                B = [435+15, 235 + 60, 615, 285 + 60]
                C = [223+15, 310 + 60, 400, 360 + 60]
                D = [435+15, 310 + 60, 615, 360 + 60]
            else:
                wen = [336, 95 + 60, 605, 228 + 60]
                A = [336+15, 275 + 60, 420, 305 + 60]
                B = [486+15, 275 + 60, 600, 305 + 60]
                C = [336+15, 355 + 60, 420, 380 + 60]
                D = [486+15, 355 + 60, 600, 380 + 60]
            self.handleed_wen=mubiao[wen[1]:wen[3],wen[0]:wen[2]]
            self.handleed_A  =  mubiao[A[1]:A[3],A[0]:A[2]]
            self.handleed_B  =  mubiao[B[1]:B[3],B[0]:B[2]]
            self.handleed_C  =  mubiao[C[1]:C[3],C[0]:C[2]]
            self.handleed_D  =  mubiao[D[1]:D[3],D[0]:D[2]]
            return 1
        except:
            return 0

    def jiancekaishi_dianshi(self, ):
        """
        检测题目中是否出现了题目
        返回:出现题目 返回1
            出错    返回0
            异常    返回2
        """
        x, y, *_ = get_screen.get_bbox()
        startimg, *_ = get_screen.get_screen(bbox=[x + 190, y + 130, x + 240, y + 175])
        startimg = cv.cvtColor(np.asarray(startimg), cv.COLOR_RGB2BGR)
        startimg37, *_ = get_screen.get_screen(bbox=[x + 305, y +100 , x + 350, y + 145])
        startimg37 = cv.cvtColor(np.asarray(startimg37), cv.COLOR_RGB2BGR)
        flag1=self.get_areaready(startimg, self.startwen)
        flag2=self.get_areaready(startimg37, self.startwen37)
        while flag1==0 and flag2==0:
            x, y, *_ = get_screen.get_bbox()
            startimg, *_ = get_screen.get_screen(bbox=[x + 190, y + 130, x + 240, y + 175])
            startimg = cv.cvtColor(np.asarray(startimg), cv.COLOR_RGB2BGR)
            startimg37, *_ = get_screen.get_screen(bbox=[x + 305, y +100 , x + 355, y + 145])
            startimg37 = cv.cvtColor(np.asarray(startimg37), cv.COLOR_RGB2BGR)
            flag1=self.get_areaready(startimg, self.startwen)
            flag2=self.get_areaready(startimg37, self.startwen37)
        self.flag37=1 if flag2 else 0
        return 1
    def jiancekaishi_dianshi2(self ):
        """
        检测题目中是否出现了题目
        返回:出现题目 返回1
            出错    返回0
            异常    返回2
        """
        x, y, *_ = get_screen.get_bbox()
        startimg, *_ = get_screen.get_screen(bbox=[x + 520, y + 440, x + 640, y + 515])
        startimg = cv.cvtColor(np.asarray(startimg), cv.COLOR_RGB2BGR)
        return self.get_areaready(startimg, self.leave)

    # 针对房都蔚
    def huida_fangduwei(self, xuangxiang):
        """
        自动回答程序
        传入:选项
        传出:成功 1
            失败 0
        """
        x, y, *_ = get_screen.get_bbox()

        if xuangxiang == 'A':
            click(200 + x, 300 + y)
        elif xuangxiang == 'B':
            click(x + 450, y + 300)
        elif xuangxiang == 'C':
            click(x + 200, y + 375)
        elif xuangxiang == 'D':
            click(x + 450, y + 375)
        time.sleep(time_click_delay)
        move(x + 290, y + 410)
        return 1

    def handleimgquestion_fangduwei(self):
        """
        针对房都蔚划分区域
        输入:numpy格式的图像
        输出:二值化后的问题,选项ABCD的图像
        """
        try:
            mubiao, _ = get_screen.get_screen()
            mubiao = cv.cvtColor(np.asarray(mubiao), cv.COLOR_RGB2BGR)
            wen = [98, 125 + 60, 570, 175 + 60]
            A = [98,  213 + 60,  310, 270 + 60]
            B = [345, 213 + 60,  570, 270 + 60]
            C = [98,  290 + 60,  310, 350 + 60]
            D = [345, 290 + 60,  570, 350 + 60]

            self.handleed_wen=mubiao[wen[1]:wen[3],wen[0]:wen[2]]
            self.handleed_A  =  mubiao[A[1]:A[3],A[0]:A[2]]
            self.handleed_B  =  mubiao[B[1]:B[3],B[0]:B[2]]
            self.handleed_C  =  mubiao[C[1]:C[3],C[0]:C[2]]
            self.handleed_D  =  mubiao[D[1]:D[3],D[0]:D[2]]
            return 1
        except:
            return 0

    def jiancekaishi_fangduwei(self ):
        """
        检测题目中是否出现了题目
        返回:出现题目 返回1
            出错    返回0
            异常    返回2
        """
        try:
            x, y, *_ = get_screen.get_bbox()
            startimg, *_ = get_screen.get_screen(bbox=[x + 65, y + 175, x + 110, y + 220])
            startimg = cv.cvtColor(np.asarray(startimg), cv.COLOR_RGB2BGR)
            # cv.imshow('1',startimg)
            # cv.waitKey(0)
            #print(self.get_areaready(startimg, self.startwen))
            return self.get_areaready(startimg, self.startwen)
        except:
            return 0
    def jiancekaishi_fangduwei2(self ):
        """
        检测题目中是否出现了题目
        返回:出现题目 返回1
            出错    返回0
            异常    返回2
        """
        try:
            x, y, *_ = get_screen.get_bbox()
            startimg, *_ = get_screen.get_screen(bbox=[x + 480, y + 410, x + 610, y + 480])
            startimg = cv.cvtColor(np.asarray(startimg), cv.COLOR_RGB2BGR)
            return self.get_areaready(startimg, self.leave)
        except:
            return 0


    def handleimgquestion_shenghuochangshi(self):
        """
        针对房都蔚划分区域
        输入:numpy格式的图像
        输出:二值化后的问题,选项ABCD的图像
        """
        try:

            mubiao, _ = get_screen.get_screen()
            mubiao = cv.cvtColor(np.asarray(mubiao), cv.COLOR_RGB2BGR)
            wen = [225, 160, 605, 240]
            A = [228, 300, 392, 340]
            B = [437, 300, 606, 340]
            C = [228,370, 392, 415]
            D = [437, 370, 606, 415]
            self.handleed_wen=mubiao[wen[1]:wen[3],wen[0]:wen[2]]
            self.handleed_A  =  mubiao[A[1]:A[3],A[0]:A[2]]
            self.handleed_B  =  mubiao[B[1]:B[3],B[0]:B[2]]
            self.handleed_C  =  mubiao[C[1]:C[3],C[0]:C[2]]
            self.handleed_D  =  mubiao[D[1]:D[3],D[0]:D[2]]

            return 1
        except:
            return 0
    def jiancekaishi_shenghuofujiati(self):
        """
        检测题目中是否出现了题目
        返回:出现题目 返回1
            出错    返回0
            异常    返回2
        """
        x, y, *_ = get_screen.get_bbox()
        startimg, *_ = get_screen.get_screen(bbox=[x + 196, y + 130, x + 242, y + 174])
        startimg = cv.cvtColor(np.asarray(startimg), cv.COLOR_RGB2BGR)
        return self.get_areaready(startimg, self.startwen)
    def jiancekaishi_shenghuofujiati2(self):
        """
        检测题目中是否出现了题目
        返回:出现题目 返回1
            出错    返回0
            异常    返回2
        """
        x, y, *_ = get_screen.get_bbox()
        startimg, *_ = get_screen.get_screen(bbox=[x + 550, y + 460, x + 640, y + 500])
        startimg = cv.cvtColor(np.asarray(startimg), cv.COLOR_RGB2BGR)
        return self.get_areaready(startimg, self.leave)
pic2txt=Pic2txt()
get_screen=Get_Screen()
findkey=Findkey()
answerapi=answerAPI()
nums=0

t=time.time()
while True:
    time_save = time.time()

    #等待题目出现阶段
    print('*' * 20 + '等待题目出现' + '*' * 20)
    while answerapi.jiancekaishi() == 0:
        pass
    while answerapi.jiancekaishi2() == 0:
        pass
    time.sleep(0.01)
    while answerapi.jiancekaishi() == 0:
        pass
    while answerapi.jiancekaishi2() == 0:
        pass

    #识别题目阶段
    print('*' * 20 + '得到题目区域' + '*' * 20)
    result = answerapi.hanleimgquestion()
    if not result:
        print("ERROR:题目获取失败")
        continue
    result = answerapi.gettxtfromAPI()
    if result==0:
        print("ERROR:文字识别失败")
        continue
    print("得到题目")
    print(f"问:{answerapi.txt_wens}")
    print(f"A:{answerapi.txt_As}")
    print(f"B:{answerapi.txt_Bs}")
    print(f"C:{answerapi.txt_Cs}")
    print(f"D:{answerapi.txt_Ds}")


    #匹配答案阶段
    result = findkey.getanswerfromquestion(str(answerapi.txt_wens), str(answerapi.txt_As), str(answerapi.txt_Bs), str(answerapi.txt_Cs), str(answerapi.txt_Ds))
    if result==0:
        print("ERROR:答案匹配失败")
        continue
    print(result)
    if result[4]<80:
        print("ERROR:答案匹配度过低重试")
        #尝试通过选项寻找答案
        result=findkey.getanswerfromkey(str(answerapi.txt_As), str(answerapi.txt_Bs), str(answerapi.txt_Cs), str(answerapi.txt_Ds))
        if result:
            print("通过选项匹配到了答案,新的答案'")
            print(result)
        else:
            continue

    #点击阶段
        # try:
        #     #获取点击前的屏幕截图
        #     x, y, *_ = get_screen.get_bbox()
        #     fist, *_ = get_screen.get_screen(bbox=[x + 230, y + 160, x + 600, y + 250])
        #     fist = cv.cvtColor(np.asarray(fist), cv.COLOR_RGB2BGR)
        #     cv.imwrite('./data/fist.png', fist)
        # except:
        #     pass
    #点击操作
    dianjiresult = answerapi.huida(result[0])

    if titlename == '梦幻西游自动答题训练器':
        x, y, *_ = get_screen.get_bbox()
        click(732+x,522+y)
    time.sleep(time_dati_delay)

    nums += 1
    print(f"单题耗时{(time.time() - time_save):.4f} 已答{nums}题累计耗时{time.time() - t:.4f}")
    print('-' * 50)
"""    #检查点击是否成功
    x, y, *_ = get_screen.get_bbox()
    seccod, *_ = get_screen.get_screen(bbox=[x + 230, y + 160, x + 600, y + 250])
    second = cv.cvtColor(np.asarray(seccod), cv.COLOR_RGB2BGR)

    # 获取点击后的屏幕截图1
    cv.imwrite('./data/second1.png', second)
    flag=answerapi.get_diff("./data/fist.png", './data/second1.png') #比较两图是否相等
    if flag==0:
        print("点击成功")
    else:
        while answerapi.jiancekaishi() == 0:
            pass
        while answerapi.jiancekaishi2() == 0:
            pass
        time.sleep(0.01)
        x, y, *_ = get_screen.get_bbox()
        seccod, *_ = get_screen.get_screen(bbox=[x + 230, y + 160, x + 600, y + 250])
        second = cv.cvtColor(np.asarray(seccod), cv.COLOR_RGB2BGR)
        cv.imwrite('./data/second2.png', second)
        if flag:
            dianjiresult = answerapi.huida(result[0])
            x, y, *_ = get_screen.get_bbox()
            seccod, *_ = get_screen.get_screen(bbox=[x + 230, y + 160, x + 600, y + 250])
            second = cv.cvtColor(np.asarray(seccod), cv.COLOR_RGB2BGR)
            flag = answerapi.get_diff("./data/fist.png", './data/second2.png')
            if flag==0:
                print("点击成功")
            else:
                print("多次点击失败,请手动点击后继续回车运行程序")
                input("\n")



    nums += 1
    print(f"单题耗时{(time.time() - time_save):.4f} 已答{nums}题累计耗时{time.time() - t:.4f}")
    print('-' * 50)
"""

